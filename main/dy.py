import openpyxl
import json
import io


def dynamodb():
    pass


def mapping_excel_to_json(excel_file, excel_sheet, json_file_name):
    book = openpyxl.load_workbook(excel_file)
    sheet = book[excel_sheet]
    max_row = sheet.max_row
    max_column = sheet.max_column
    result = []
    heads = []
    for column in range(max_column):
        heads.append(sheet.cell(1, column + 1).value)
    for row in range(max_row):
        if row == 0:
            continue
        one_line = {}
        for column in range(max_column):
            k = heads[column]
            cell = sheet.cell(row + 1, column + 1)
            value = cell.value
            one_line[k] = str(value)
        print(one_line)
        result.append(one_line)
    book.close()

    file = io.open(json_file_name, 'w', encoding='utf-8')
    txt = json.dumps(result, indent=2, ensure_ascii=False)
    file.write(txt)
    file.close()
    return file

def mapping_env_tran(source_env, target_env, path):
    pass
